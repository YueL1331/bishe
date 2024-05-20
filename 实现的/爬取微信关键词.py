import requests
import pandas as pd
import time
import warnings
from urllib3.exceptions import InsecureRequestWarning
from openpyxl import Workbook, load_workbook

# 禁用未验证的 HTTPS 请求警告
warnings.simplefilter('ignore', InsecureRequestWarning)

# 文件路径
file_path = '/Users/a58/Documents/微信关键词爬取/爬取关键词.xlsx'
output_excel_path = '/Users/a58/Documents/微信关键词爬取/爬取到的数据.xlsx'

# 读取Excel文件
df = pd.read_excel(file_path)

# 提取关键词（跳过第一行标题）
keywords = df.iloc[:, 0].dropna().tolist() + df.iloc[:, 1].dropna().tolist()

# 用户指定的开始关键词
start_keyword = ''  # 设置为 '' 时，从第一个词开始查询；否则填写关键词

# 查找开始关键词的索引
start_index = 0
if start_keyword:
    try:
        start_index = keywords.index(start_keyword) + 1
    except ValueError:
        print(f"关键词 '{start_keyword}' 不存在于关键词列表中。将从第一个关键词开始。")

# 请求数据的headers
headers = {
    'Host': 'search.weixin.qq.com',
    'xweb_xhr': '1',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/107.0.0.0 Safari/537.36 MicroMessenger/6.8.0(0x16080000) NetType/WIFI MiniProgramEnv/Mac '
                  'MacWechat/WMPF MacWechat/3.8.5(0x1308050a)XWEB/31023',
    'Content-Type': 'application/json',
    'Accept': '*/*',
    'Sec-Fetch-Site': 'cross-site',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Dest': 'empty',
    'Referer': 'https://servicewechat.com/wxc026e7662ec26a3a/53/page-frame.html',
    'Accept-Language': 'zh-CN,zh;q=0.9',
}

# 读取或创建工作表
try:
    book = load_workbook(output_excel_path)
    sheet = book.active
except FileNotFoundError:
    book = Workbook()
    sheet = book.active
    sheet.title = 'Sheet1'
    book.save(output_excel_path)
    book = load_workbook(output_excel_path)
    sheet = book['Sheet1']

# 读取已处理的关键词
processed_keywords = set()
for row in sheet.iter_rows(values_only=True):
    if row[0]:  # 第一列是关键词
        processed_keywords.add(row[0])

# 初始化一个标志，表示是否已经写入时间行
time_row_written = any(row[0] == '' for row in sheet.iter_rows(values_only=True))

# 从指定索引开始处理关键词
for keyword in keywords[start_index:]:
    if keyword in processed_keywords:
        continue  # 跳过已处理的关键词

    json_data = {
        'openid': 'ov4ns0OQxlddUfLF76GZKF5hoq5o',
        'search_key': '1716184851258760_1666778467',
        'cgi_name': 'GetDefaultIndex',
        'start_ymd': '20240510',
        'end_ymd': '20240516',
        'query': keyword,
    }

    try:
        response = requests.post('https://search.weixin.qq.com/cgi-bin/wxaweb/wxindex', headers=headers, json=json_data, verify=False)
        response.raise_for_status()
        response_data = response.json()

        # 检查反扒返回
        if response_data.get('code') == -10000:
            print("Auth failed. Terminating the script.")
            break

        resp_list = response_data.get('content', {}).get('resp_list', [])
        if resp_list:
            query = resp_list[0]['query']
            time_scores = resp_list[0]['indexes'][0]['time_indexes']

            # 初始化存储数据
            if not time_row_written:
                time_row = ['']  # 第一个单元格留空
                for item in time_scores:
                    time_row.append(item['time'])
                sheet.append(time_row)
                time_row_written = True  # 设置标志，表示时间行已写入

            score_row = [query]  # 第一行是关键词

            # 填充数据
            for item in time_scores:
                score_row.append(item['score'])

            print(f"Parsed data for query '{query}':")
            print("Score row:", score_row)

            # 将数据追加到工作表中
            sheet.append(score_row)
            book.save(output_excel_path)
        else:
            # 如果没有数据，插入包含 "--" 的行
            print(f"No data found for keyword '{keyword}'. Adding '--' as placeholder.")
            if not time_row_written:
                time_scores_length = 1  # 如果还没有时间行，设定默认长度为1
            else:
                # 获取时间行长度
                time_scores_length = len(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))) - 1
            no_data_row = [keyword] + ['--'] * time_scores_length
            sheet.append(no_data_row)
            book.save(output_excel_path)

    except Exception as e:
        print(f"Error processing keyword '{keyword}': {e}")
        if not time_row_written:
            time_scores_length = 1
        else:
            # 获取已写入时间行的长度
            time_scores_length = len(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))) - 1
        no_data_row = [keyword] + ['--'] * time_scores_length
        sheet.append(no_data_row)
        book.save(output_excel_path)

    # 固定间隔5秒，模拟用户操作
    time.sleep(5)

# 保存工作簿
book.save(output_excel_path)
book.close()
