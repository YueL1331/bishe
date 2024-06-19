import hashlib
import time

import requests
import json
import pandas as pd
import urllib3
from openpyxl import Workbook, load_workbook

# 忽略 SSL 警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

headers = {
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Cache-Control': 'no-cache',
    'Connection': 'keep-alive',
    'Content-Type': 'application/json',
    'Origin': 'https://www.tianyancha.com',
    'Pragma': 'no-cache',
    'Referer': 'https://www.tianyancha.com/',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-site',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
    'X-TYCID': '8029b170296011efa426e1981bd8dd18',
    'sec-ch-ua': '"Google Chrome";v="125", "Chromium";v="125", "Not.A/Brand";v="24"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"macOS"',
}

# 设置代理
proxies = {
    'http': 'http://t15873905246814:gg6xwmok@u273.kdltps.com:15818',
    'https': 'http://t15873905246814:gg6xwmok@u273.kdltps.com:15818',
}


def search_mind(keyword: str):
    params = {
        '_': str(int(round(time.time() * 1000))),
    }

    json_data = {
        'keyword': keyword,
    }

    response = requests.post(
        'https://capi.tianyancha.com/cloud-tempest/search/suggest/v5',
        params=params,
        headers=headers,
        json=json_data,
        proxies=proxies,
        verify=False
    )
    if response.status_code == 200:
        return response.text
    else:
        return {}


def batch_search():
    df = pd.read_excel('../data/qcc查询query.xlsx')
    keywords = df['关键词']
    total = len(keywords)
    xlsx_path = "../data/tyc_company_id_mapping.xlsx"
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['序号', '企业ID', '企业名称', '查询关键字', '响应报文'])
        wb.save(xlsx_path)
    last_row = ws.max_row
    try:
        start_index = int(ws[last_row][0].value)
    except Exception:
        start_index = 0
    for index in range(total):
        if index < start_index:
            continue
        key = keywords[index]
        company_list = ''
        keyNo = ''
        name = ''
        try:
            company_list = search_mind(key)
            print(company_list)
            company_data = json.loads(company_list)
            first_company = company_data.get('data')[0]
            keyNo = str(first_company.get('id', ''))
            name = str(first_company.get('comName', ''))
            print("完成第" + str(index + 1) + key+":\t"+str(keyNo)+"\t"+str(name))
        except Exception as e3:
            print("完成第" + str(index + 1) + key + ":异常:Exception")
        finally:
            try:
                ws.append([index + 1, keyNo, name, key, company_list])
            except ValueError:
                print(index + 1, keyNo, name, key, company_list)
                ws.append([index + 1, '', '', key, ''])
            if index % 100 == 0:
                wb.save(xlsx_path)
    wb.save(xlsx_path)


if __name__ == '__main__':
    batch_search()
