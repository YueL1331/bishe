import requests
import openpyxl
import os
import time
import json
import datetime

# 忽略 SSL 警告
requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)

# 设置代理
proxies = {
    'http': 'http://t15873905246814:gg6xwmok@u273.kdltps.com:15818',
    'https': 'http://t15873905246814:gg6xwmok@u273.kdltps.com:15818',
}

# 加载要处理的企业ID
keyNo_wb = openpyxl.load_workbook('company_id_mapping 2.xlsx')
keyNo_ws = keyNo_wb.active

# 读取 keyNo 列中的值
keyNos = []
start_index = 0  # 读取的起始序号
for row in keyNo_ws.iter_rows(min_row=2):
    keyNos.append({
        "序号": row[0].value,
        "企业ID": row[1].value,
        "企业名称": row[2].value,
        "查询关键字": row[3].value
    })

# 创建 Excel 文件
output_file_count = 1
output_filename = f"qcc爬取/企业信息{output_file_count}.xlsx"
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "企业信息"

headers_list = [
    "序号", "企业ID", "企业名称", "查询关键字", "统一社会信用代码", "注册号", "类型（企业类型）", "组成形式", "登记机关",
    "经营场所/住所（公司地址）", "经营范围", "名称/企业名称", "经营者/法定代表人", "注册日期/成立日期",
    "核准日期", "登记状态", "注册资本", "营业期限自", "营业期限至", "股东出资信息", "主要人员信息",
    "分支机构信息", "行业", "电话"
]

sheet.append(headers_list)
workbook.save(output_filename)

# 获取目标页面内容并写入 Excel
for idx, keyNo_info in enumerate(keyNos[start_index:], start=start_index + 1):
    keyNo = keyNo_info["企业ID"]
    seq_no = keyNo_info["序号"]
    company_name = keyNo_info["企业名称"]
    query_keyword = keyNo_info["查询关键字"]

    # 如果企业ID为空，写入前四列后跳过
    if not keyNo:
        sheet.append([seq_no, keyNo, company_name, query_keyword] + [""] * (len(headers_list) - 4))
        workbook.save(output_filename)
        continue

    # 更新时间戳
    current_timestamp = int(datetime.datetime.now().timestamp() * 1000)

    # 设置请求头
    headers = {
        'Host': 'xcx.qcc.com',
        'x-request-device-type': 'Android',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36 MicroMessenger/6.8.0(0x16080000) NetType/WIFI MiniProgramEnv/Mac MacWechat/WMPF MacWechat/3.8.7(0x13080712) XWEB/1191',
        'content-type': 'application/json',
        'qcc-version': '1.0.0',
        'authmini': 'Bearer d317ba05a53546ec7ea2d10d65ae206f',
        'xweb_xhr': '1',
        'xcx-version': '2024.06.15',
        'qcc-platform': 'mp-weixin',
        'qcc-currentpage': '/company-subpackages/business/index',
        'qcc-timestamp': str(current_timestamp),
        'qcc-refpage': '/company-subpackages/detail/index',
        'accept': '*/*',
        'sec-fetch-site': 'cross-site',
        'sec-fetch-mode': 'cors',
        'sec-fetch-dest': 'empty',
        'referer': 'https://servicewechat.com/wx395200814fcd7599/353/page-frame.html',
        'accept-language': 'zh-CN,zh;q=0.9',
    }

    params = {
        'token': 'd317ba05a53546ec7ea2d10d65ae206f',
        't': str(current_timestamp),
        'unique': keyNo,
    }

    target_url = f"https://xcx.qcc.com/mp-weixin/forwardApp/v6/base/getEntDetail"
    print(f"Processing {target_url} for keyNo {keyNo}")

    try:
        response = requests.get(target_url, params=params, headers=headers, proxies=proxies, verify=False, timeout=10)
        if response.status_code != 200:
            print(f"Failed to retrieve the target page for {keyNo}. Status code: {response.status_code}")
            print(f"Previous sequence number was {seq_no - 1}")
            break
    except requests.RequestException as e:
        print(f"Error retrieving {target_url} for keyNo {keyNo}: {e}")
        print(f"Previous sequence number was {seq_no - 1}")
        break

    # 解析JSON
    data = response.json()
    print(f"Received data for keyNo {keyNo}: {json.dumps(data, ensure_ascii=False, indent=4)}")  # 添加日志打印完整的返回内容

    # 提取所需字段
    company_info = data.get("result", {}).get("Company", {})
    contact_info = data.get("result", {}).get("ContactInfo", {})
    employees = data.get("result", {}).get("Employees", [])
    partners = data.get("result", {}).get("Partners", [])

    # print(f"Company info: {company_info}")  # 打印公司信息
    print(f"Contact info: {contact_info}")  # 打印联系方式信息
    print(f"Employees: {employees}")  # 打印主要人员信息
    print(f"Partners: {partners}")  # 打印股东信息

    # 准备行数据
    row_data = [
        seq_no,
        keyNo,
        company_name,
        query_keyword,
        company_info.get("CreditCode", ""),
        company_info.get("OrgNo", ""),
        company_info.get("EconKind", ""),
        "",  # 组成形式，页面中没有找到
        company_info.get("BelongOrg", ""),
        company_info.get("Address", ""),
        company_info.get("Scope", ""),
        company_info.get("Name", ""),
        company_info.get("Oper", {}).get("Name", ""),
        company_info.get("StartDate", ""),
        company_info.get("CheckDate", ""),
        company_info.get("Status", ""),
        company_info.get("RegistCapi", ""),
        company_info.get("TermStart", ""),
        company_info.get("TeamEnd", ""),
        "",  # 股东出资信息，后面处理
        "",  # 主要人员信息，后面处理
        "",  # 分支机构信息，不在JSON数据中
        company_info.get("IndustryV3", {}).get("SubIndustry", ""),
        contact_info.get("PhoneNumber", "")
    ]

    # 处理股东出资信息
    shareholders_info = []
    for partner in partners:
        shareholders_info.append({
            "股东名称": partner.get("StockName", ""),
            "持股比例": partner.get("StockPercent", ""),
            "认缴出资额": partner.get("ShouldCapi", ""),
            "认缴出资日期": partner.get("ShoudDate", "")
        })
    row_data[19] = str(shareholders_info) if shareholders_info else ""

    # 处理主要人员信息
    key_personnel_info = []
    for idx, employee in enumerate(employees, start=1):
        key_personnel_info.append({
            "序号": idx,
            "姓名": employee.get("Name", ""),
            "职务": employee.get("Job", "")
        })
    row_data[20] = str(key_personnel_info) if key_personnel_info else ""

    # 写入 Excel
    sheet.append(row_data)
    workbook.save(output_filename)
    for header, value in zip(headers_list, row_data):
        print(f"{header}: {value}")

    # 每处理1000个就创建一个新文件
    if (idx + 1) % 1000 == 0:
        output_file_count += 1
        output_filename = f"qcc爬取/企业信息{output_file_count}.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "企业信息"
        sheet.append(headers_list)
        workbook.save(output_filename)

    time.sleep(3)  # 避免请求过于频繁

print("处理完成")
