import requests
from bs4 import BeautifulSoup
import urllib3
import openpyxl
import os
import time

# 忽略 SSL 警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 设置请求头
headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
}

# 设置获取到的Cookies
cookies = {
    'QCCSESSID': 'c1f36f78d2d41f8c36cc98e3f9',
    'qcc_did': 'daf3e340-610c-412d-b313-dee14e652b4c',
    'UM_distinctid': '18ff0b5f620604-0a0a706b91bacd-1b525637-13c680-18ff0b5f6212167',
    'tfstk': 'f9e2klN2nA45dxViz4MZ8IoFJwHx1vQBQRgsjlqicq0DhKTgalaThfbxhRzrH8NXCq_xQl4IQw_Cd9ZYDxHGRwtwVFGSIcXsSLs93fHtIw_SgJVd2AUIhbtA_ujrfcgMnPmmE40-bmAgSqckrDoKIV4iIgWojcvMmcAmE3Y83J29zcGcxzSx-UdRJj03mqqbc8cpR2qmzdviumlVao0yId2zwoFmw2Wyr4Zgv5czQtpjBuNU754ljd0aZDzEAr6k3AVuzJl32wJxu74YQxGdQhozTu443A5eD0w4qJhaHZJouzEuOxehydz3VkwxwR7erVFjvYm4aGAoofjPMFnls8eTgFAZi0nrRgSuELZOvYQF9yA96jIx42smiCdti0p6gJ3p6ChAD0uCmjf..',
    'acw_tc': '2f624a7417183485899127487e13026f738f273465107728fef64c51e783e3',
    'CNZZDATA1254842228': '1490140565-1717730408-%7C1718349198'
}

# 设置代理
proxies = {
    'http': 'http://t15873905246814:gg6xwmok@u273.kdltps.com:15818',
    'https': 'http://t15873905246814:gg6xwmok@u273.kdltps.com:15818',
}

# 创建输出文件夹
output_dir = "qcc爬取"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# 加载要处理的企业ID
keyNo_wb = openpyxl.load_workbook('company_id_mapping 2.xlsx')
keyNo_ws = keyNo_wb.active

# 读取 keyNo 列中的值
keyNos = []
start_index = 0  # 读取的起始序号
for row in keyNo_ws.iter_rows(min_row=2):
    keyNos.append({
        "序号": row[0].value,
        "企业ID": row[1].value,
        "企业名称": row[2].value,
        "查询关键字": row[3].value
    })

# 创建 Excel 文件
output_file_count = 1
output_filename = os.path.join(output_dir, f"企业信息{output_file_count}.xlsx")
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "企业信息"

headers_list = [
    "序号", "企业ID", "企业名称", "查询关键字", "统一社会信用代码", "注册号", "类型（企业类型）", "组成形式", "登记机关",
    "经营场所/住所（公司地址）", "经营范围", "名称/企业名称", "经营者/法定代表人", "注册日期/成立日期",
    "核准日期", "登记状态", "注册资本", "营业期限自", "营业期限至", "股东出资信息", "主要人员信息",
    "分支机构信息", "行业", "电话"
]

sheet.append(headers_list)
workbook.save(output_filename)

last_successful_seq_no = None

# 获取目标页面内容并写入 Excel
for idx, keyNo_info in enumerate(keyNos[start_index:], start=start_index + 1):
    keyNo = keyNo_info["企业ID"]
    seq_no = keyNo_info["序号"]
    company_name = keyNo_info["企业名称"]
    query_keyword = keyNo_info["查询关键字"]

    # 如果企业ID为空，写入前四列后跳过
    if not keyNo:
        row_data = [seq_no, keyNo, company_name, query_keyword] + [""] * (len(headers_list) - 4)
        sheet.append(row_data)
        workbook.save(output_filename)
        for header, value in zip(headers_list, row_data):
            print(f"{header}: {value}")
        last_successful_seq_no = seq_no
        continue

    target_url = f"https://www.qcc.com/firm/{keyNo}.html"
    print(f"Processing {target_url}")

    try:
        response = requests.get(target_url, headers=headers, cookies=cookies, proxies=proxies, verify=False, timeout=10)
        if response.status_code == 423:
            print(f"Failed to retrieve {target_url}: Status code {response.status_code}")
            print(f"Last successful seq no: {last_successful_seq_no}")
            break
        elif response.status_code != 200:
            print(f"Failed to retrieve {target_url}: Status code {response.status_code}")
            row_data = [seq_no, keyNo, company_name, query_keyword] + [""] * (len(headers_list) - 4)
            sheet.append(row_data)
            workbook.save(output_filename)
            for header, value in zip(headers_list, row_data):
                print(f"{header}: {value}")
            last_successful_seq_no = seq_no
            continue
    except requests.RequestException as e:
        print(f"Error retrieving {target_url}: {e}")
        row_data = [seq_no, keyNo, company_name, query_keyword] + [""] * (len(headers_list) - 4)
        sheet.append(row_data)
        workbook.save(output_filename)
        for header, value in zip(headers_list, row_data):
            print(f"{header}: {value}")
        last_successful_seq_no = seq_no
        continue

    # 解析HTML
    soup = BeautifulSoup(response.content, 'html.parser')

    data = {}

    # 提取工商信息
    cominfo_section = soup.find('section', {'id': 'cominfo'})
    company_info = {}  # 默认情况下为空字典
    if cominfo_section:
        rows = cominfo_section.find_all('tr')

        for row in rows:
            cells = row.find_all('td')
            if len(cells) >= 2:
                label_1 = cells[0].get_text(strip=True)
                value_1 = cells[1].get_text(strip=True)

                # 特别处理法定代表人字段
                if '法定代表人' in label_1:
                    name_tag = cells[1].find('a')
                    if name_tag:
                        value_1 = name_tag.get_text(strip=True)

                company_info[label_1] = value_1

                if len(cells) >= 4:
                    label_2 = cells[2].get_text(strip=True)
                    value_2 = cells[3].get_text(strip=True)
                    company_info[label_2] = value_2

                if len(cells) >= 6:
                    label_3 = cells[4].get_text(strip=True)
                    value_3 = cells[5].get_text(strip=True)
                    company_info[label_3] = value_3

        data['工商信息'] = company_info

    # 提取股东信息
    partner_section = soup.find('div', {'class': 'app-tree-table'})
    if partner_section:
        rows = partner_section.find_all('tr')
        partner_info = []

        for row in rows[1:]:  # 跳过表头
            cells = row.find_all('td')
            if len(cells) >= 2:
                name_span = cells[1].find('span', {'class': 'name'})
                partner_name = name_span.get_text(strip=True) if name_span else cells[1].get_text(strip=True)
                partner = {
                    '序号': cells[0].get_text(strip=True),
                    '股东名称': partner_name,
                    '持股比例': cells[2].get_text(strip=True),
                    '认缴出资额': cells[3].get_text(strip=True),
                }
                partner_info.append(partner)

        data['股东信息'] = partner_info

    # 提取主要人员信息
    mainmember_section = soup.find('section', {'id': 'mainmember'})
    if mainmember_section:
        rows = mainmember_section.find_all('tr')
        mainmember_info = []

        for row in rows[1:]:  # 跳过表头
            cells = row.find_all('td')
            if len(cells) >= 3:
                name_span = cells[1].find('span', {'class': 'name'})
                member_name = name_span.get_text(strip=True) if name_span else cells[1].get_text(strip=True)
                member = {
                    '序号': cells[0].get_text(strip=True),
                    '姓名': member_name,
                    '职务': cells[2].get_text(strip=True),
                }
                mainmember_info.append(member)

        data['主要人员信息'] = mainmember_info

    # 提取其他相关信息
    header_info_part = soup.find('div', {'class': 'ui-header-info-part'})
    phone = ""
    if header_info_part:
        rlines = header_info_part.find_all('div', {'class': 'rline'})

        for rline in rlines:
            label_span = rline.find('span', {'class': 'f'})
            value_span = rline.find('span', {'class': 'val'})
            if label_span and value_span:
                label = label_span.get_text(strip=True)
                value = value_span.get_text(strip=True)
                if "电话" in label:
                    phone = value

    # 提取分支机构信息
    branchelist_section = soup.find('section', {'id': 'branchelist'})
    if branchelist_section:
        rows = branchelist_section.find_all('tr')
        branchelist_info = []

        for row in rows[1:]:  # 跳过表头
            cells = row.find_all('td')
            if len(cells) >= 6:
                # 检查是否包含不需要的标签
                if not (cells[1].find('span', {'class': 'app-auto-logo'}) or
                        cells[1].find('div', {'class': 'app-tdcoy-tags app-tags margin-type-default'}) or
                        cells[1].find('span', {'class': 'tail-tag'})):
                    name_span = cells[1].find('span', {'class': 'name'})
                    company_name = name_span.get_text(strip=True) if name_span else cells[1].get_text(strip=True)
                    name_span = cells[2].find('span', {'class': 'name'})
                    person_in_charge = name_span.get_text(strip=True) if name_span else ""
                    branch = {
                        '序号': cells[0].get_text(strip=True),
                        '企业名称': company_name,
                        '负责人': person_in_charge,
                        '地区': cells[3].get_text(strip=True),
                        '成立日期': cells[4].get_text(strip=True),
                        '状态': cells[5].get_text(strip=True),
                    }
                    branchelist_info.append(branch)

        data['分支机构信息'] = branchelist_info

    # 处理营业期限
    operating_period = company_info.get('营业期限', '')
    period_parts = operating_period.split('至')
    if len(period_parts) == 2:
        business_info_start = period_parts[0].strip()
        business_info_end = period_parts[1].strip()
    else:
        business_info_start = ""
        business_info_end = ""

    # 获取需要的信息
    business_info = data.get('工商信息', {})
    shareholders_info = data.get('股东信息', [])
    key_personnel_info = data.get('主要人员信息', [])
    branch_info = data.get('分支机构信息', [])

    row_data = [
        seq_no,
        keyNo,
        company_name,
        query_keyword,
        business_info.get("统一社会信用代码", ""),
        business_info.get("工商注册号", ""),
        business_info.get("企业类型", ""),
        "",  # 组成形式，页面中没有找到
        business_info.get("登记机关", ""),
        business_info.get("注册地址", ""),
        business_info.get("经营范围", ""),
        business_info.get("企业名称", ""),
        business_info.get("法定代表人", ""),
        business_info.get("成立日期", ""),
        business_info.get("核准日期", ""),
        business_info.get("登记状态", ""),
        business_info.get("注册资本", ""),
        business_info_start,  # 营业期限自
        business_info_end,  # 营业期限至
        str(shareholders_info) if shareholders_info else "",
        str(key_personnel_info) if key_personnel_info else "",
        str(branch_info) if branch_info else "",
        business_info.get("国标行业", ""),
        phone
    ]

    sheet.append(row_data)
    workbook.save(output_filename)

    # 在终端输出写入的信息
    for header, value in zip(headers_list, row_data):
        print(f"{header}: {value}")

    last_successful_seq_no = seq_no

    # 每处理1000个就创建一个新文件
    if (idx + 1) % 1000 == 0:
        output_file_count += 1
        output_filename = os.path.join(output_dir, f"企业信息{output_file_count}.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "企业信息"
        sheet.append(headers_list)
        workbook.save(output_filename)

    time.sleep(3)  # 避免请求过于频繁

print("处理完成")
