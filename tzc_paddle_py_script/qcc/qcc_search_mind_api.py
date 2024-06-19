import hashlib
import time

import requests
from urllib import parse
import json
import hmac
import pandas as pd
from openpyxl import Workbook, load_workbook
import random


def search_mind(keyword: str):
    cookies = {
        'QCCSESSID': 'c6f68701cd4246adc2a500b650',
        'qcc_did': 'bf37a478-7d83-4e54-b0e2-c845d2c4e57b',
        'UM_distinctid': '1900a8306fb7b0-0ecc7ea4c8cc72-1a525637-13c680-1900a8306fc8fc',
        'tfstk': 'fups1vZDrEpF8LKCo1nFFFazXeXXhFMzlosvqneaDOBTkrKR8GJZSnQXle737ClDuEaX7e8ficbVkZT2DCoEUY-MjtXvl4krUAfgtxv1krWAx5OOZ4urUvWaVJ9KzIRw0oWCmwIOX5FtA6I5XtUOXEQd9gIzWtBvkHnCbgyYBtQYJwIJPNe1Yc_HfDwG9cR7xMxOR8tHdGI_shQQH-pB41_-hwwYHps6xLWf7RH5JCRvQ9ptJzXpqB8OATMbcMOWvOI6oY2GB3KpM19jgubDOHdPOBqZjM96VKs59yn1od86h1v-zJb9OpvfQBu_UZfvTdfkQ2e1eQxNQILKD-QX9iIyn875MhNbA_2fAaoIASV2EFvGlCxYuIfOxMuqADNym1IhAMmIASVc6MjedDiQanf..',
        'CNZZDATA1254842228': '992078368-1718163278-https%253A%252F%252Fwww.qcc.com%252F%7C1718185629',
    }

    headers = {
        # '1feb93e28350392deb7e': 'd191641105660514177966decd2cd42a1f8f09cd56d421b3d8be466a833b08a02b6e13a934e78b18c3f6e2f4b261c5dbe9737a8eb4a3615c83e58b6d81566fe9',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'zh-CN,zh;q=0.9',
        'cache-control': 'no-cache',
        # 'cookie': 'QCCSESSID=c6f68701cd4246adc2a500b650; qcc_did=bf37a478-7d83-4e54-b0e2-c845d2c4e57b; UM_distinctid=1900a8306fb7b0-0ecc7ea4c8cc72-1a525637-13c680-1900a8306fc8fc; tfstk=fups1vZDrEpF8LKCo1nFFFazXeXXhFMzlosvqneaDOBTkrKR8GJZSnQXle737ClDuEaX7e8ficbVkZT2DCoEUY-MjtXvl4krUAfgtxv1krWAx5OOZ4urUvWaVJ9KzIRw0oWCmwIOX5FtA6I5XtUOXEQd9gIzWtBvkHnCbgyYBtQYJwIJPNe1Yc_HfDwG9cR7xMxOR8tHdGI_shQQH-pB41_-hwwYHps6xLWf7RH5JCRvQ9ptJzXpqB8OATMbcMOWvOI6oY2GB3KpM19jgubDOHdPOBqZjM96VKs59yn1od86h1v-zJb9OpvfQBu_UZfvTdfkQ2e1eQxNQILKD-QX9iIyn875MhNbA_2fAaoIASV2EFvGlCxYuIfOxMuqADNym1IhAMmIASVc6MjedDiQanf..; CNZZDATA1254842228=992078368-1718163278-https%253A%252F%252Fwww.qcc.com%252F%7C1718185629',
        'pragma': 'no-cache',
        'priority': 'u=1, i',
        'referer': 'https://www.qcc.com/',
        'sec-ch-ua': '"Google Chrome";v="125", "Chromium";v="125", "Not.A/Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"macOS"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
        'x-pid': '83501dfb338244157a6af1e1380ff0dd',
        'x-requested-with': 'XMLHttpRequest',
    }

    params = {
        'mindKeyWords': 'true',
        'mindType': '9',
        'pageSize': '5',
        'person': 'true',
        'searchKey': keyword,
        'suggest': 'true',
    }

    header = dynamic_header(params)
    headers[header[0]] = header[1]
    '''proxies = {
        'http': 'http://O23071010551916725899:pwd=c0WBPdyW&pid=-1&cid=-1&sip=0&dip=0&uid=@proxy-58.vpsnb.net:14223',
        'https': 'http://O23071010551916725899:pwd=c0WBPdyW&pid=-1&cid=-1&sip=0&dip=0&uid=@proxy-58.vpsnb.net:14223',
    }'''
    proxies = {
        'http': 'http://t15873905246814:gg6xwmok@u273.kdltps.com:15818',
        'https': 'http://t15873905246814:gg6xwmok@u273.kdltps.com:15818'
    }
    response = requests.get('https://www.qcc.com/api/search/searchMind', params=params, cookies=cookies,
                            headers=headers, proxies=proxies)
    if response.status_code == 200:
        return response.text
    else:
        return {}


def dynamic_header(params: dict):
    e = {
        "url": "/api/search/searchMind",
        "baseURL": "https://www.qcc.com",
        "params": params,
        "data": {}
    }
    t = e['url'].replace(e['baseURL'], "")
    n = None
    if e['params']:
        n = parse.urlencode(e['params'])
    else:
        n = parse.urlencode({})
    if n:
        if t.find('?') == -1:
            t += ('?' + n)
        else:
            t += ('&' + n)
    t = t.lower()
    name = gen_name(t, e['data'])
    # print('name=', name)
    windowTid = 'c400b081f47505d5f59e7f2b09813034'
    val = gen_val(t, e['data'], windowTid)
    # print("val=", val)
    return name, val


def gen_name(t: str, e: dict):
    if t == '' or t is None:
        t = '/'
    t = t.lower()
    if e is None:
        e = {}
    n = json.dumps(e).lower()
    url = encode_url(t)
    n_ = (t + n)
    hmac_md5 = hmac.new(url.encode('utf-8'), n_.encode('utf-8'), hashlib.sha512)
    digest = hmac_md5.hexdigest()
    # print(len(digest))
    # print(digest)
    return digest.lower()[8:28]


def gen_val(n: str, e: dict, t: str):
    if n == '' or n is None:
        n = '/'
    n = n.lower()
    if e is None:
        e = {}
    if t is None:
        t = ''
    i = json.dumps(e).lower()
    hmac_md5 = hmac.new(encode_url(n).encode('utf-8'), (n + "pathString" + i + t).encode('utf-8'), hashlib.sha512)
    return hmac_md5.hexdigest()


codes = {0: 'W', 1: 'l', 2: 'k', 3: 'B', 4: 'Q', 5: 'g', 6: 'f', 7: 'i', 8: 'i', 9: 'r', 10: 'v', 11: '6', 12: 'A',
         13: 'K', 14: 'N', 15: 'k', 16: '4', 17: 'L', 18: '1', 19: '8'}
le = len(codes)


def encode_url(e: str):
    t = e + e
    n = ''
    for i in range(len(t)):
        a = ord(t[i]) % le
        n += codes[a]
    return n


def batch_search():
    df = pd.read_excel('qcc查询query.xlsx')
    keywords = df['关键词']
    total = len(keywords)
    try:
        wb = load_workbook("../data/company_id_mapping.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['序号', '企业ID', '企业名称', '查询关键字', '响应报文'])
        wb.save("company_id_mapping.xlsx")
    last_row = ws.max_row
    try:
        start_index = ws[last_row][0].value
    except Exception:
        start_index = 0
    for index in range(total):
        if index < start_index:
            continue
        key = keywords[index]
        company_list = ''
        keyNo = ''
        name = ''
        try:
            company_list = search_mind(key)
            print(company_list)
            company_data = json.loads(company_list)
            first_company = company_data['list'][0]
            keyNo = first_company['KeyNo']
            name = first_company['name']
            print("完成第" + str(index + 1) + key)
        except ValueError as e1:
            print("完成第" + str(index + 1) + key + ":异常:ValueError")
        except IndexError as e2:
            print("完成第" + str(index + 1) + key + ":异常:IndexError")
        except Exception as e3:
            print("完成第" + str(index + 1) + key + ":异常:Exception")
        finally:
            try:
                ws.append([index + 1, keyNo, name, key, company_list])
            except ValueError:
                print(index + 1, keyNo, name, key, company_list)
                ws.append([index + 1, '', '', key, ''])
            if index % 100 == 0:
                wb.save("company_id_mapping.xlsx")
    wb.save("company_id_mapping.xlsx")


if __name__ == '__main__':
    batch_search()
    # print(dynamic_header('厦门巨龙'))
