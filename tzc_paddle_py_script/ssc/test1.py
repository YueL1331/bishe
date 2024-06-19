import re
import pandas as pd
import requests
from bs4 import BeautifulSoup
import os
import time
from openpyxl import load_workbook

# 填入实际的 cookies 和 headers
cookies = {
    'your_cookie_name': 'your_cookie_value',
    # 其他 cookies
}

headers = {
    'User-Agent': 'your_user_agent',
    'Accept': 'your_accept_header',
    # 其他 headers
}


def fetch_company_info(company_id, cookies, headers):
    url = f'https://www.sscha.com/detail/{company_id}'
    print(f"Fetching info for company ID: {company_id}")  # Debugging line
    print(f"Headers: {headers}")  # Debugging line
    response = requests.get(url, cookies=cookies, headers=headers, verify=False)

    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        data = {
            '统一社会信用代码': '',
            '注册号': '',
            '类型': '',
            '登记机关': '',
            '经营场所/住所': '',
            '经营范围': '',
            '名称/企业名称': '',
            '经营者/法定代表人': '',
            '注册日期/成立日期': '',
            '核准日期': '',
            '登记状态': '',
            '注册资本': '',
            '营业期限自': '',
            '营业期限至': '',
            '股东出资信息': [],
            '主要人员信息': [],
            '分支机构信息': [],
            '行业': '',
            '电话': ''
        }

        # 提取公司基本信息
        basic_info = soup.find('div', {'class': 'un-business-info'})
        if basic_info:
            spans = basic_info.find_all('span')
            spans_text = [span.text.strip() for span in spans]
            if '企业名称' in spans_text:
                name_index = spans_text.index('企业名称') + 1
                data['名称/企业名称'] = spans[name_index].text.strip()
            if '统一社会信用代码' in spans_text:
                code_index = spans_text.index('统一社会信用代码') + 1
                data['统一社会信用代码'] = spans[code_index].text.strip()
            if '企业类型' in spans_text:
                type_index = spans_text.index('企业类型') + 1
                data['类型'] = spans[type_index].text.strip()
            if '登记机关' in spans_text:
                agency_index = spans_text.index('登记机关') + 1
                data['登记机关'] = spans[agency_index].text.strip()
            if '注册地址' in spans_text:
                address_index = spans_text.index('注册地址') + 1
                data['经营场所/住所'] = spans[address_index].text.strip()
            if '经营范围' in spans_text:
                scope_index = spans_text.index('经营范围') + 1
                data['经营范围'] = spans[scope_index].text.strip()
            if '法定代表人' in spans_text:
                representative_index = spans_text.index('法定代表人') + 1
                data['经营者/法定代表人'] = spans[representative_index].find('a').text.strip()
            if '成立日期' in spans_text:
                registration_date_index = spans_text.index('成立日期') + 1
                data['注册日期/成立日期'] = spans[registration_date_index].text.strip()
            if '核准日期' in spans_text:
                approval_date_index = spans_text.index('核准日期') + 1
                data['核准日期'] = spans[approval_date_index].text.strip()
            if '登记状态' in spans_text:
                status_index = spans_text.index('登记状态') + 1
                data['登记状态'] = spans[status_index].text.strip()
            if '注册资本' in spans_text:
                capital_index = spans_text.index('注册资本') + 1
                data['注册资本'] = spans[capital_index].text.strip()
            if '营业期限' in spans_text:
                period_index = spans_text.index('营业期限') + 1
                period = spans[period_index].text.strip().split(' 至 ')
                data['营业期限自'] = period[0] if len(period) > 1 else ''
                data['营业期限至'] = period[1] if len(period) > 1 else ''
            if '所属行业' in spans_text:
                industry_index = spans_text.index('所属行业') + 1
                data['行业'] = spans[industry_index].text.strip()
            if '工商注册号' in spans_text:
                registration_number_index = spans_text.index('工商注册号') + 1
                data['注册号'] = spans[registration_number_index].text.strip()
            if '联系电话' in spans_text:
                phone_index = spans_text.index('联系电话') + 1
                data['电话'] = spans[phone_index].text.strip()

        # 获取股东出资信息
        shareholders_table = soup.find('div', class_='un-shareholder')
        if shareholders_table:
            shareholders_rows = shareholders_table.find_all('tr')[1:]  # 跳过标题行
            for row in shareholders_rows:
                cells = row.find_all('td')
                shareholder = {}
                if len(cells) > 0 and cells[0].text.strip():
                    shareholder['股东姓名'] = cells[0].text.strip()
                if len(cells) > 1 and cells[1].text.strip():
                    shareholder['股东类型'] = cells[1].text.strip()
                if len(cells) > 2 and cells[2].text.strip():
                    shareholder['持股比例'] = cells[2].text.strip()
                if len(cells) > 3 and cells[3].text.strip():
                    shareholder['认缴出资额'] = cells[3].text.strip()
                if len(cells) > 4 and cells[4].text.strip():
                    shareholder['认缴出资时间'] = cells[4].text.strip()

                if shareholder:
                    data['股东出资信息'].append(shareholder)

        # 提取主要人员信息
        senior_staff = soup.find('div', {'class': 'un-senior'})
        if senior_staff:
            rows = senior_staff.find_all('tr')
            headers = [header.text.strip() for header in rows[0].find_all('th')]
            # 确定 "职务" 或 "职位" 列的位置
            position_index = -1
            if '职务' in headers:
                position_index = headers.index('职务')
            elif '职位' in headers:
                position_index = headers.index('职位')

            for row in rows[1:]:
                cols = row.find_all('td')
                if len(cols) >= 2:  # 确保每一行有正确的列数
                    staff_info = {
                        '姓名': cols[0].text.strip(),
                        '职务': cols[position_index].text.strip() if position_index != -1 else '',
                    }
                    data['主要人员信息'].append(staff_info)

        # 提取分支机构信息
        branches = soup.find('div', {'class': 'un-branch'})
        if branches:
            rows = branches.find_all('tr')[1:]
            for row in rows:
                cols = row.find_all('td')
                if len(cols) == 5:  # 确保每一行有正确的列数
                    branch_info = {
                        '企业名称': cols[0].text.strip(),
                        '企业状态': cols[1].text.strip(),
                        '法定代表人': cols[2].text.strip(),
                        '地区': cols[3].text.strip(),
                        '成立时间': cols[4].text.strip(),
                    }
                    data['分支机构信息'].append(branch_info)

        return data
    else:
        return None


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, **to_excel_kwargs):
    if not os.path.isfile(filename):
        df.to_excel(filename, sheet_name=sheet_name, startrow=startrow if startrow is not None else 0, index=False,
                    **to_excel_kwargs)
        return

    with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        writer.book = load_workbook(filename)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, header=False, index=False, **to_excel_kwargs)


def main():
    input_file = 'company_id_mapping 2.xlsx'
    output_folder = 'output'
    os.makedirs(output_folder, exist_ok=True)

    df = pd.read_excel(input_file)
    start_sequence = 1  # 设置导入序号的起始值
    file_base_name = 1  # 文件序号的起始值
    file_name = os.path.join(output_folder, f'ssc{file_base_name}.xlsx')

    for index, row in df.iterrows():
        company_id = row.iloc[1]  # 修改为使用 .iloc
        info = fetch_company_info(company_id, cookies, headers)  # 传递 cookies 和 headers

        if info:
            record = {
                '序号': start_sequence + index,
                '企业ID': company_id,
                '企业名称': row.iloc[2],  # 修改为使用 .iloc
                '查询关键词': row.iloc[3],  # 修改为使用 .iloc
                **info
            }

            # 打印写入内容
            print(f"写入内容: {record}")

            # 每查询到一条数据就写入到 Excel 文件中
            append_df_to_excel(file_name, pd.DataFrame([record]))

            time.sleep(1)  # 每个爬取之间暂停1秒


if __name__ == "__main__":
    main()
