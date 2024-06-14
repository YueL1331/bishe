import requests
from bs4 import BeautifulSoup
import urllib3
import openpyxl
import os

# 忽略 SSL 警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 设置请求头
headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
}

# 设置获取到的Cookies
cookies = {
    'QCCSESSID': '988f17dd434c091be146b78d76',
    'qcc_did': 'daf3e340-610c-412d-b313-dee14e652b4c',
    'UM_distinctid': '18ff0b5f620604-0a0a706b91bacd-1b525637-13c680-18ff0b5f6212167',
    'tfstk': 'fTnq6j0VmlcSrugga7ZaTSl_vxZY5kdBbcN_IADghSVcci1izAGtcPAYcclrKXJTfoiXbVrxVqs_GIEZIXZwOBtBAxHYDlABOx7BWYrgIhNMsLd5GlEMOQaqx2p8XXM1HnLZELy_C52isPflq7ecslV0S_buK7VgjlcgqT2gL-jcmiDlEUHjxcCzCrvnpczgjNA3oWD0ixMjqJ5QtxVPjGV43rJrn7SGj0u-Hq48T35uNADKSzowDiZtF2lZSbvGj5uuLXyIG3j4j4crYyDJTiFqrjnacyKNQ8uEEjo4beWYFmkovPovggVqwY0QDDARbXDSdmU-bB5u6qwLq-uyIaV0SgRA68Yi2c3VsNz0e8PBULmlVApL2B8KtNQTkKezOK4cWNU0eG_qvWQOWrpYUW90o',
    'acw_tc': '707c9fce17182588583674650e415548d862f2099ce35fc6adb6d686e7df38',
    'CNZZDATA1254842228': '1490140565-1717730408-%7C1718259007'
}

# 创建一个会话
session = requests.Session()

# 设置Cookies
session.cookies.update(cookies)

# 读取keyNo.xlsx文件
keyNo_file = 'keyNo.xlsx'
keyNo_wb = openpyxl.load_workbook(keyNo_file)
keyNo_ws = keyNo_wb.active

# 获取所有keyNo
keyNos = [row[0] for row in keyNo_ws.iter_rows(min_row=2, values_only=True)]

# 创建目录和保存Excel文件
output_dir = "qcc爬取"
output_file = os.path.join(output_dir, "企业信息.xlsx")
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# 创建Excel并写入数据
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "企业信息"

headers_list = [
    "企业ID", "企业名称", "统一社会信用代码", "注册号", "类型（企业类型）", "组成形式", "登记机关",
    "经营场所/住所（公司地址）", "经营范围", "名称/企业名称", "经营者/法定代表人",
    "注册日期/成立日期", "核准日期", "登记状态", "注册资本", "营业期限自", "营业期限至",
    "股东出资信息", "主要人员信息", "分支机构信息", "行业", "电话"
]

sheet.append(headers_list)

# 处理每个keyNo
for keyNo in keyNos:
    target_url = f"https://www.qcc.com/firm/{keyNo}.html"

    # 获取目标页面内容
    response = session.get(target_url, headers=headers, verify=False)
    if response.status_code == 200:
        # 解析HTML
        soup = BeautifulSoup(response.content, 'html.parser')

        data = {}

        # 提取工商信息
        cominfo_section = soup.find('section', {'id': 'cominfo'})
        if cominfo_section:
            rows = cominfo_section.find_all('tr')
            company_info = {}

            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    label_1 = cells[0].get_text(strip=True)
                    value_1 = cells[1].get_text(strip=True)

                    # 特别处理法定代表人字段
                    if '法定代表人' in label_1:
                        name_tag = cells[1].find('a')
                        if name_tag:
                            value_1 = name_tag.get_text(strip=True)

                    company_info[label_1] = value_1

                    if len(cells) >= 4:
                        label_2 = cells[2].get_text(strip=True)
                        value_2 = cells[3].get_text(strip=True)
                        company_info[label_2] = value_2

                    if len(cells) >= 6:
                        label_3 = cells[4].get_text(strip=True)
                        value_3 = cells[5].get_text(strip=True)
                        company_info[label_3] = value_3

            data['工商信息'] = company_info

        # 提取股东信息
        partner_section = soup.find('div', {'class': 'app-tree-table'})
        if partner_section:
            rows = partner_section.find_all('tr')
            partner_info = []

            for row in rows[1:]:  # 跳过表头
                cells = row.find_all('td')
                if len(cells) >= 4:
                    name_span = cells[1].find('span', {'class': 'name'})
                    partner_name = name_span.get_text(strip=True) if name_span else cells[1].get_text(strip=True)
                    partner = {
                        '序号': cells[0].get_text(strip=True),
                        '股东名称': partner_name,
                        '持股比例': cells[2].get_text(strip=True),
                        '认缴出资额': cells[3].get_text(strip=True),
                        '认缴出资日期': cells[4].get_text(strip=True),
                    }
                    partner_info.append(partner)

            data['股东信息'] = partner_info

        # 提取主要人员信息
        mainmember_section = soup.find('section', {'id': 'mainmember'})
        if mainmember_section:
            rows = mainmember_section.find_all('tr')
            mainmember_info = []

            for row in rows[1:]:  # 跳过表头
                cells = row.find_all('td')
                if len(cells) >= 3:
                    name_span = cells[1].find('span', {'class': 'name'})
                    member_name = name_span.get_text(strip=True) if name_span else cells[1].get_text(strip=True)
                    member = {
                        '序号': cells[0].get_text(strip=True),
                        '姓名': member_name,
                        '职务': cells[2].get_text(strip=True),
                    }
                    mainmember_info.append(member)

            data['主要人员信息'] = mainmember_info

        # 提取其他相关信息
        header_info_part = soup.find('div', {'class': 'ui-header-info-part'})
        if header_info_part:
            header_info = {}
            rlines = header_info_part.find_all('div', {'class': 'rline'})

            for rline in rlines:
                label_span = rline.find('span', {'class': 'f'})
                value_span = rline.find('span', {'class': 'val'})
                if label_span and value_span:
                    label = label_span.get_text(strip=True)
                    value = value_span.get_text(strip=True)
                    if "电话" in label:
                        header_info["电话"] = value

            data['其他信息'] = header_info

        # 提取分支机构信息
        branchelist_section = soup.find('section', {'id': 'branchelist'})
        if branchelist_section:
            rows = branchelist_section.find_all('tr')
            branchelist_info = []

            for row in rows[1:]:  # 跳过表头
                cells = row.find_all('td')
                if len(cells) >= 6:
                    # 检查是否包含不需要的标签
                    if not (cells[1].find('span', {'class': 'app-auto-logo'}) or
                            cells[1].find('div', {'class': 'app-tdcoy-tags app-tags margin-type-default'}) or
                            cells[1].find('span', {'class': 'tail-tag'})):
                        name_span = cells[1].find('span', {'class': 'name'})
                        company_name = name_span.get_text(strip=True) if name_span else cells[1].get_text(strip=True)
                        name_span = cells[2].find('span', {'class': 'name'})
                        person_in_charge = name_span.get_text(strip=True) if name_span else None
                        branch = {
                            '序号': cells[0].get_text(strip=True),
                            '企业名称': company_name,
                            '负责人': person_in_charge,
                            '地区': cells[3].get_text(strip=True),
                            '成立日期': cells[4].get_text(strip=True),
                            '状态': cells[5].get_text(strip=True),
                        }
                        branchelist_info.append(branch)

            data['分支机构信息'] = branchelist_info

        # 处理营业期限
        operating_period = company_info.get('营业期限', '')
        period_parts = operating_period.split('至')
        if len(period_parts) == 2:
            business_info_start = period_parts[0].strip()
            business_info_end = period_parts[1].strip()
        else:
            business_info_start = "null"
            business_info_end = "null"

        # 获取需要的信息
        business_info = data.get('工商信息', {})
        shareholders_info = data.get('股东信息', [])
        key_personnel_info = data.get('主要人员信息', [])
        branch_info = data.get('分支机构信息', [])
        other_info = data.get('其他信息', {})

        row_data = [
            keyNo,
            business_info.get("企业名称", "null"),
            business_info.get("统一社会信用代码", "null"),
            business_info.get("工商注册号", "null"),
            business_info.get("企业类型", "null"),
            "null",  # 组成形式，页面中没有找到
            business_info.get("登记机关", "null"),
            business_info.get("注册地址", "null"),
            business_info.get("经营范围", "null"),
            business_info.get("企业名称", "null"),
            business_info.get("法定代表人", "null"),
            business_info.get("成立日期", "null"),
            business_info.get("核准日期", "null"),
            business_info.get("登记状态", "null"),
            business_info.get("注册资本", "null"),
            business_info_start,  # 营业期限自
            business_info_end,  # 营业期限至
            str(shareholders_info) if shareholders_info else "null",
            str(key_personnel_info) if key_personnel_info else "null",
            str(branch_info) if branch_info else "null",
            business_info.get("国标行业", "null"),
            other_info.get("电话", "null")
        ]

        sheet.append(row_data)

    else:
        print(f"Failed to retrieve the target page for {keyNo}. Status code: {response.status_code}")

# 保存Excel文件
workbook.save(output_file)
print(f"数据已写入到 {output_file} 文件中。")

# 打印终端输出
for row in sheet.iter_rows(min_row=2, values_only=True):
    for header, value in zip(headers_list, row):
        print(f"{header}: {value}")
